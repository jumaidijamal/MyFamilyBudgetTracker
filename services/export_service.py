import csv

from datetime import date
from openpyxl import Workbook


class ExportService:

    def __init__(
        self,
        sheet_service
    ):
        self.sheet_service = (
            sheet_service
        )

    # ==========================
    # GET MONTH TRANSACTIONS
    # ==========================

    def get_month_transactions(self):

        sheet = (
            self.sheet_service
            .get_sheet(
                "Transactions"
            )
        )

        rows = (
            sheet
            .get_all_records()
        )

        today = date.today()

        export_rows = []

        for row in rows:

            if (
                row["Status"]
                == "Deleted"
            ):
                continue

            trx_date = str(
                row[
                    "TransactionDate"
                ]
            )

            if not trx_date.startswith(
                f"{today.year}-"
                f"{today.month:02d}"
            ):
                continue

            export_rows.append(
                row
            )

        return export_rows

    # ==========================
    # EXPORT CSV
    # ==========================

    def export_month_csv(self):

        rows = (
            self
            .get_month_transactions()
        )

        today = date.today()

        filename = (
            f"Transactions_"
            f"{today.year}_"
            f"{today.month:02d}.csv"
        )

        with open(
            filename,
            "w",
            newline="",
            encoding="utf-8-sig"
        ) as file:

            if rows:

                writer = (
                    csv.DictWriter(
                        file,
                        fieldnames=(
                            rows[0].keys()
                        )
                    )
                )

                writer.writeheader()
                writer.writerows(
                    rows
                )

        return filename

    # ==========================
    # EXPORT XLSX
    # ==========================

    def export_month_xlsx(self):

        rows = (
            self
            .get_month_transactions()
        )

        wallets = (
            self.sheet_service
            .get_sheet(
                "Wallets"
            )
            .get_all_records()
        )

        wb = Workbook()

        ws_trx = wb.active
        ws_trx.title = (
            "Transactions"
        )

        ws_summary = (
            wb.create_sheet(
                "Summary"
            )
        )

        ws_wallet = (
            wb.create_sheet(
                "Wallets"
            )
        )

        # ==========================
        # SHEET TRANSACTIONS
        # ==========================

        ws_trx.append(
            [
                "Date",
                "Type",
                "Category",
                "Wallet",
                "Amount",
                "Description"
            ]
        )

        income = 0
        expense = 0

        for row in rows:

            ws_trx.append(
                [
                    row[
                        "TransactionDate"
                    ],

                    row[
                        "Type"
                    ],

                    row[
                        "CategoryName"
                    ],

                    row[
                        "WalletName"
                    ],

                    row[
                        "Amount"
                    ],

                    row[
                        "Description"
                    ]
                ]
            )

            amount = float(
                row[
                    "Amount"
                ]
            )

            if (
                row["Type"]
                == "income"
            ):

                income += amount

            else:

                expense += amount

        # ==========================
        # SHEET SUMMARY
        # ==========================

        net = (
            income
            - expense
        )

        ws_summary.append(
            [
                "Item",
                "Amount"
            ]
        )

        ws_summary.append(
            [
                "Income",
                income
            ]
        )

        ws_summary.append(
            [
                "Expense",
                expense
            ]
        )

        ws_summary.append(
            [
                "Net",
                net
            ]
        )

        # ==========================
        # SHEET WALLETS
        # ==========================

        ws_wallet.append(
            [
                "Wallet",
                "Balance"
            ]
        )

        for wallet in wallets:

            ws_wallet.append(
                [
                    wallet[
                        "WalletName"
                    ],

                    wallet[
                        "CurrentBalance"
                    ]
                ]
            )

        # ==========================
        # SAVE FILE
        # ==========================

        today = date.today()

        filename = (
            f"Transactions_"
            f"{today.year}_"
            f"{today.month:02d}.xlsx"
        )

        wb.save(
            filename
        )

        return filename